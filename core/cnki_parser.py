from dataclasses import dataclass
from typing import List
import sys
import os
import openpyxl
import pandas as pd
from collections import defaultdict


class CNKIFormatItem(object):
    def __init__(self):
        self.elements = []


@dataclass
class Article(object):
    elements: List[str]
    # RT:str
    # SR:str
    # A1:str
    # AD:str
    # T1:str
    # TF:str
    # YR:str
    # IS:str
    # vo:str
    # OP:str
    # K1:str
    # AB:str
    # SN:str
    # CN:str
    # LA:str
    # DS:str


@dataclass
class Author(object):
    name: str
    articles: List[Article]
    article_count: int


def format_parser_helper(path, format):
    try:
        files = os.listdir(path)
        for file in files:
            if format == 'CNKI':
                parser = CNKIFormatParser()
                parser.parse_cnki_files(file)
    except Exception as e:
        print(str(e))


class SciMetricsAnalyzer(object):
    def __init__(self, format, data_path, project_path):
        self.format = format
        self.data_path = data_path
        self.project_path = project_path
        self.articles = []
        self.analyze_articles()
        self.co_authors = {}
        self.first_authors = {}

    def analyze_articles(self):
        try:
            # files = os.listdir(self.data_path)
            # for file in files:
            if self.format == 'CNKI':
                parser = CNKIFormatParser()
                self.articles = parser.parse_cnki_files(self.data_path)
        except Exception as e:
            print(str(e))
        print('{}-{}'.format('文章总数：', len(self.articles)))

    def analyze_authors(self):
        try:
            if self.format == 'CNKI':
                parser = CNKIFormatParser()
                return parser.analyze_authors(self.articles)
        except Exception as e:
            print(str(e))



    # def analyze_authors(self):
    #     for article in self.articles:
    #         try:
    #             str_authors = article.elements[2]
    #             str_authors_list = str_authors.rstrip(';').split(';')
    #             # 将第一作者和其他共同作者区分开
    #             # 第一作者处理
    #             first_author = str_authors_list[0]
    #             if first_author in self.first_authors:
    #                 self.first_authors[first_author].articles.append(article)
    #                 self.first_authors[first_author].article_count += 1
    #             else:
    #                 author = Author(first_author, [], 0)
    #                 author.articles.append(article)
    #                 author.article_count = 1
    #                 self.first_authors[first_author] = author
    #
    #             for str_author in str_authors_list:
    #                 if str_author in self.co_authors:
    #                     self.co_authors[str_author].articles.append(article)
    #                     self.co_authors[str_author].article_count += 1
    #                 else:
    #                     author = Author(str_author, [], 0)
    #                     author.articles.append(article)
    #                     author.article_count = 1
    #                     self.co_authors[str_author] = author
    #         except Exception as e:
    #             print(str(e))
    #     # self.first_authors = sorted(self.first_authors.items(), key=lambda x: x[1].article_count, reverse=True)
    #     print('OK')
        



class CNKIFormatParser(object):
    def __init__(self):
        pass

    def parse_cnki_files(self, file):
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            rows = ws.max_row   
            columns = ws.max_column         
            cnki_items = []
            chinese = True
            for row in range(2, rows+1):
                guard = ws.cell(row, 1).value
                if guard == None:
                    continue
                elif guard == 'RT':
                    chinese = False
                    continue
                if chinese is True:
                    item = Article([])
                    for column in range(1, columns):                    
                        item.elements.append(ws.cell(row=row, column=column).value)
                elif chinese is False:
                    item = Article([])
                    for column in range(1, 10):                    
                        item.elements.append(ws.cell(row=row, column=column).value)
                    item.elements.append(None)
                    item.elements.append(ws.cell(row=row, column=10).value)
                    item.elements.append(ws.cell(row=row, column=11).value)  
                    item.elements.append(None)
                    item.elements.append(None)
                    item.elements.append(None)
                    item.elements.append(ws.cell(row=row, column=12).value)
                cnki_items.append(item)
            return cnki_items
        except Exception as e:
            print(str(e))

    def analyze_authors(self, articles):
        authors_stat = defaultdict(dict)
        for article in articles:
            try:
                author_in_article = article.elements[2]
                authors_list = author_in_article.rstrip(';').split(';')
                authors_list = list(filter(None, authors_list))
                # 区分第一作者和共同作者，将第一作者和共同作者的频次独立处理
                first_author = authors_list[0]
                if first_author in authors_stat and 'articleAsFirst' in authors_stat[first_author]:
                    authors_stat[first_author]['articleAsFirst'] += 1
                else:
                    authors_stat[first_author]['articleAsFirst'] = 1
                    authors_stat[first_author]['articleAsCo'] = 0
                    authors_stat[first_author]['CoOccur'] = defaultdict(dict)
                # 统计共同作者的文章频次，并且构建与第一作者的共现矩阵
                # 一般来说，只统计第一作者与共同作者之间的共现，不统计共同作者之间的共现
                for co_author in authors_list[1:]:
                    if co_author in authors_stat and 'articleAsCo' in authors_stat[co_author]:
                        authors_stat[co_author]['articleAsCo'] += 1
                    else:
                        authors_stat[co_author]['articleAsCo'] = 1
                        authors_stat[co_author]['articleAsFirst'] = 0
                        authors_stat[co_author]['CoOccur'] = defaultdict(dict)
                    A, B = first_author, co_author
                    # 固定共现序列，以比较序小的在前
                    if A > B :
                        A, B = B, A
                    if B in authors_stat[A]['CoOccur']:
                        authors_stat[A]['CoOccur'][B] += 1
                    else:
                        authors_stat[A]['CoOccur'][B] = 1
            except Exception as e:
                pass
        return authors_stat

# if __name__ == '__main__':
# path = "F:\\002-测试数据\\NewNetworkAttack\\input\\CNKI-637293898682050000.xlsx"
# analyzer = SciMetricsAnalyzer('CNKI', path, None)
# analyzer.analyze_authors()

# authors = {}

# authors['A'] = 1
# authors['B'] = 2
# authors['C'] = 7
# authors['D'] = 4
# authors = sorted(authors.items(), key=lambda x: x[1], reverse=True)

# print(authors)
